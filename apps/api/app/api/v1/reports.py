import io
from typing import Annotated

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from app.api.deps import get_db, require_viewer
from app.repositories.execution_log_repository import ExecutionLogRepository
from app.repositories.patch_repository import PatchRepository
from app.schemas.auth import UserResponse
from app.schemas.report import ReportItem

router = APIRouter()


def _build_report_items(db: Session) -> list[ReportItem]:
    repository = ExecutionLogRepository(db)
    logs = repository.list_recent(limit=1000)
    patch_category: dict[str, str] = {
        p.id: p.category for p in PatchRepository(db).list_all()
    }
    return [
        ReportItem(
            date=log.executed_at.strftime("%d/%m %H:%M"),
            schedule=log.schedule_name,
            machine=log.machine_name,
            patch=log.patch_id,
            platform=log.platform,
            severity=log.severity,
            category=patch_category.get(log.patch_id, "unknown"),
            result=log.result,
            duration=f"{log.duration_seconds // 60}m {log.duration_seconds % 60:02d}s",
        )
        for log in logs
    ]


@router.get("", response_model=list[ReportItem])
def list_reports(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[UserResponse, Depends(require_viewer)],
) -> list[ReportItem]:
    return _build_report_items(db)


@router.get("/export.xlsx")
def export_xlsx(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[UserResponse, Depends(require_viewer)],
) -> StreamingResponse:
    """Generate a real .xlsx file using openpyxl and stream it to the client."""
    items = _build_report_items(db)

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório operacional"

    # --- Header row styling ---
    header_fill = PatternFill(start_color="172033", end_color="172033", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = ["Data", "Janela", "Máquina", "Patch", "Plataforma", "Severidade", "Categoria", "Resultado", "Duração"]
    column_widths = [16, 28, 24, 32, 14, 14, 14, 14, 12]

    for col_idx, (header, width) in enumerate(zip(headers, column_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 22

    # --- Data rows ---
    result_colors = {
        "applied": "D1FAE5",
        "completed": "D1FAE5",
        "failed": "FEE2E2",
        "skipped": "FEF9C3",
    }

    for row_idx, item in enumerate(items, start=2):
        row_data = [
            item.date,
            item.schedule,
            item.machine,
            item.patch,
            item.platform,
            item.severity,
            item.category,
            item.result,
            item.duration,
        ]
        result_color = result_colors.get(item.result)
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center")
            if result_color:
                cell.fill = PatternFill(
                    start_color=result_color, end_color=result_color, fill_type="solid"
                )

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter on header
    ws.auto_filter.ref = ws.dimensions

    # --- Stream response ---
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    from datetime import datetime
    filename = f"patch-manager-relatorio-{datetime.now().strftime('%Y-%m-%d')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
